from openpyxl import load_workbook

PATTERN = r'.*(\d.|\d)\d\d\d\d\d(RP|B|S|BE|SE)'

if __name__ == '__main__':

    wb = load_workbook('./downloads/Остатки.xlsx')
    sheet = wb[wb.sheetnames[0]]
    wb2 = load_workbook('./downloads/global.xlsx')
    sheet2 = wb2[wb.sheetnames[0]]
    while True:
        search = input('Артикул >')
        if str(search) == 'exit':
            break
        for item in sheet.values:
            if str(item[0]).find(str(search)) != -1:
                print(f'Артикул {item[0]} в количестве {item[-1]} на складе ЭПАРХ')
        for item in sheet2.values:
            if str(item[2]).replace('.', '').find(str(search)) != -1:
                print(f'Артикул {item[2]} в количестве '
                      f'{str(item[-1]).replace(" ", "")} на '
                      f'складе '
                      f'Глобал')



